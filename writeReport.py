import csv
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, colors

from sys import argv
from os import remove


### process arguments :

parser = argparse.ArgumentParser()
parser.add_argument("csvfile",help="csv file to convert to excel")
parser.add_argument("-t","--title",help="report title",default="report")
parser.add_argument("-o","--outfile",help="output xls file name",default="special-authority-report.xlsx")

args = parser.parse_args()

input_file = args.csvfile
report_title = args.title
out_file = args.outfile

### workbook operations : 

# create workbook
wb = Workbook()

# assign active workbook
ws = wb.active

# set title
ws.title = report_title

#remove('/home/abhisek/development/python/savedSheetfromcsv.xlsx')

### open csv file for processing :

with open(input_file,'rb') as csvfile:

	out = csv.DictReader(csvfile)

	## get field attibutes
	fields = out.fieldnames
	length = len(fields)  

	## write headers
	for i in range(0,length) :
		ws.cell(row=1,column=i+1).value = fields[i]

	## format headers
	for i in range(0,length) :
		ws.cell(row=1,column=i+1).font = Font(color=colors.DARKBLUE)
	
	## write rows into spreadsheet
	for idx,row in enumerate(out,start=2):
		for col in range(0,length):
			ws.cell(row=idx,column=col+1).value = row[fields[col]]

	## format worksheet 

	column_widths = []
	for row in out:
	    for i, cell in enumerate(row):
		if len(column_widths) > i:
		    if len(cell) > column_widths[i]:
		        column_widths[i] = len(cell)
		else:
		    column_widths += [len(cell)]

	for i, column_width in enumerate(column_widths):
	    ws.column_dimensions[get_column_letter(i+1)].width = column_width

### save workbook :

wb.save( filename = out_file)
